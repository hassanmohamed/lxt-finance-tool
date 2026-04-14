"""
LXT Financial Consolidated Report — Streamlit App
===================================================
Password-protected web app that extracts General Ledger data
from 9 QuickBooks Online companies and produces a downloadable
consolidated Excel report.
"""

import base64
import html
import io
import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import bcrypt
import logging

logger = logging.getLogger(__name__)

import pandas as pd
import requests
import streamlit as st

# ─────────────────────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="LXT Financial Consolidated Report",
    page_icon="assets/lxt_logo.png",
    layout="wide",
)

# ─────────────────────────────────────────────────────────────
# LXT Logo (base64 for inline embedding)
# ─────────────────────────────────────────────────────────────
_logo_path = Path(__file__).parent / "assets" / "lxt_logo.png"
_logo_b64 = ""
if _logo_path.exists():
    _logo_b64 = base64.b64encode(_logo_path.read_bytes()).decode()

# ─────────────────────────────────────────────────────────────
# Custom CSS — LXT Branding
# ─────────────────────────────────────────────────────────────
_LXT_CSS = """
<style>
/* ══════════════════════════════════════════════════════ */
/*  Google Fonts                                          */
/* ══════════════════════════════════════════════════════ */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ══════════════════════════════════════════════════════ */
/*  CSS Variables (LXT Brand)                             */
/* ══════════════════════════════════════════════════════ */
:root {
    --lxt-orange: #FE6F38;
    --lxt-orange-light: #FF8F5E;
    --lxt-orange-dark: #E5592A;
    --lxt-navy: #001C2B;
    --lxt-navy-light: #0A1628;
    --lxt-navy-mid: #132238;
    --lxt-teal: #5E8B95;
    --lxt-teal-light: #7BA8B2;
    --lxt-white: #F0F2F6;
    --lxt-grey: #8899A6;
    --lxt-dark-text: #1A1A1A;
    --lxt-body-text: #4B4B4B;
    --lxt-card-bg: rgba(19, 34, 56, 0.6);
    --lxt-glass: rgba(19, 34, 56, 0.4);
    --lxt-border: rgba(94, 139, 149, 0.2);
    --lxt-glow: rgba(254, 111, 56, 0.3);
}

/* ══════════════════════════════════════════════════════ */
/*  Keyframe Animations                                   */
/* ══════════════════════════════════════════════════════ */
@keyframes fadeInUp {
    from { opacity: 0; transform: translateY(30px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes fadeIn {
    from { opacity: 0; }
    to   { opacity: 1; }
}
@keyframes shimmer {
    0%   { background-position: -200% 0; }
    100% { background-position: 200% 0; }
}
@keyframes gradientMove {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}
@keyframes pulse {
    0%, 100% { box-shadow: 0 0 0 0 rgba(254, 111, 56, 0.4); }
    50%      { box-shadow: 0 0 0 8px rgba(254, 111, 56, 0); }
}
@keyframes borderGlow {
    0%, 100% { border-color: rgba(94, 139, 149, 0.2); }
    50%      { border-color: rgba(254, 111, 56, 0.4); }
}
@keyframes slideInLeft {
    from { opacity: 0; transform: translateX(-20px); }
    to   { opacity: 1; transform: translateX(0); }
}

/* ══════════════════════════════════════════════════════ */
/*  Global Overrides                                      */
/* ══════════════════════════════════════════════════════ */
html, body, [class*="st-"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}
body {
    scroll-behavior: smooth;
}

/* Hide default Streamlit branding */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header[data-testid="stHeader"] {
    background: linear-gradient(180deg, var(--lxt-navy) 0%, transparent 100%) !important;
    backdrop-filter: blur(10px);
}

/* ══════════════════════════════════════════════════════ */
/*  Sidebar Styling                                       */
/* ══════════════════════════════════════════════════════ */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0D1B2A 0%, #132238 50%, #0A1628 100%) !important;
    border-right: 1px solid var(--lxt-border) !important;
}
section[data-testid="stSidebar"] .stMarkdown h3 {
    color: var(--lxt-orange) !important;
    font-weight: 600;
    letter-spacing: 0.5px;
}
section[data-testid="stSidebar"] .stCaption {
    color: var(--lxt-grey) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Buttons                                               */
/* ══════════════════════════════════════════════════════ */
.stButton > button[kind="primary"],
button[data-testid="stBaseButton-primary"] {
    background: linear-gradient(135deg, var(--lxt-orange) 0%, var(--lxt-orange-dark) 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 50px !important;
    padding: 0.6rem 2rem !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    letter-spacing: 0.3px;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    box-shadow: 0 4px 15px rgba(254, 111, 56, 0.25) !important;
}
.stButton > button[kind="primary"]:hover,
button[data-testid="stBaseButton-primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 25px rgba(254, 111, 56, 0.4) !important;
    background: linear-gradient(135deg, var(--lxt-orange-light) 0%, var(--lxt-orange) 100%) !important;
}
.stButton > button[kind="primary"]:active,
button[data-testid="stBaseButton-primary"]:active {
    transform: translateY(0) !important;
}

/* Secondary / default buttons */
.stButton > button[kind="secondary"],
button[data-testid="stBaseButton-secondary"] {
    border: 1.5px solid var(--lxt-teal) !important;
    border-radius: 50px !important;
    color: var(--lxt-teal-light) !important;
    background: transparent !important;
    transition: all 0.3s ease !important;
    font-weight: 500 !important;
}
.stButton > button[kind="secondary"]:hover,
button[data-testid="stBaseButton-secondary"]:hover {
    background: rgba(94, 139, 149, 0.1) !important;
    border-color: var(--lxt-orange) !important;
    color: var(--lxt-orange) !important;
}

/* Download button */
button[data-testid="stDownloadButton"] > button,
.stDownloadButton > button {
    background: linear-gradient(135deg, var(--lxt-orange) 0%, var(--lxt-orange-dark) 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 50px !important;
    font-weight: 600 !important;
    transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    box-shadow: 0 4px 15px rgba(254, 111, 56, 0.25) !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-2px) scale(1.02) !important;
    box-shadow: 0 8px 25px rgba(254, 111, 56, 0.4) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Expanders                                             */
/* ══════════════════════════════════════════════════════ */
.streamlit-expanderHeader {
    font-weight: 600 !important;
    font-size: 1rem !important;
    color: var(--lxt-white) !important;
    transition: color 0.3s ease !important;
}
.streamlit-expanderHeader:hover {
    color: var(--lxt-orange) !important;
}
/* Fix Material Icon text rendering as raw text in expander headers */
span[data-testid="stIconMaterial"] {
    font-size: 0 !important;
    width: 0 !important;
    height: 0 !important;
    overflow: hidden !important;
    display: none !important;
}
div[data-testid="stExpander"] {
    border: 1px solid var(--lxt-border) !important;
    border-radius: 12px !important;
    overflow: hidden;
    background: var(--lxt-card-bg) !important;
    backdrop-filter: blur(10px);
    transition: all 0.3s ease !important;
    animation: fadeInUp 0.5s ease-out;
}
div[data-testid="stExpander"]:hover {
    border-color: rgba(254, 111, 56, 0.3) !important;
    box-shadow: 0 4px 20px rgba(254, 111, 56, 0.08) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Metrics / KPI Cards                                   */
/* ══════════════════════════════════════════════════════ */
div[data-testid="stMetric"] {
    background: var(--lxt-card-bg) !important;
    backdrop-filter: blur(12px) !important;
    border: 1px solid var(--lxt-border) !important;
    border-radius: 14px !important;
    padding: 1.2rem 1.5rem !important;
    animation: fadeInUp 0.6s ease-out !important;
    transition: all 0.3s ease !important;
}
div[data-testid="stMetric"]:hover {
    border-color: var(--lxt-orange) !important;
    box-shadow: 0 0 25px rgba(254, 111, 56, 0.12) !important;
    transform: translateY(-3px);
}
div[data-testid="stMetric"] label {
    color: var(--lxt-grey) !important;
    font-size: 0.85rem !important;
    text-transform: uppercase !important;
    letter-spacing: 1px !important;
}
div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: var(--lxt-orange) !important;
    font-weight: 700 !important;
    font-size: 1.8rem !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Data Tables                                           */
/* ══════════════════════════════════════════════════════ */
.stDataFrame {
    border-radius: 12px !important;
    overflow: hidden !important;
    animation: fadeIn 0.6s ease-out;
}

/* ══════════════════════════════════════════════════════ */
/*  File Uploader                                         */
/* ══════════════════════════════════════════════════════ */
div[data-testid="stFileUploader"] {
    animation: fadeInUp 0.5s ease-out;
}
div[data-testid="stFileUploader"] section {
    border: 2px dashed var(--lxt-border) !important;
    border-radius: 12px !important;
    transition: all 0.3s ease !important;
}
div[data-testid="stFileUploader"] section:hover {
    border-color: var(--lxt-orange) !important;
    background: rgba(254, 111, 56, 0.03) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Progress Bar                                          */
/* ══════════════════════════════════════════════════════ */
.stProgress > div > div > div {
    background: linear-gradient(90deg, var(--lxt-orange), var(--lxt-orange-light)) !important;
    border-radius: 10px !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Status Container                                      */
/* ══════════════════════════════════════════════════════ */
details[data-testid="stStatusWidget"] {
    border: 1px solid var(--lxt-border) !important;
    border-radius: 12px !important;
    background: var(--lxt-card-bg) !important;
    backdrop-filter: blur(10px);
}

/* ══════════════════════════════════════════════════════ */
/*  Text Inputs                                           */
/* ══════════════════════════════════════════════════════ */
.stTextInput > div > div {
    border-radius: 12px !important;
    border-color: var(--lxt-border) !important;
    transition: all 0.3s ease !important;
}
.stTextInput > div > div:focus-within {
    border-color: var(--lxt-orange) !important;
    box-shadow: 0 0 0 3px rgba(254, 111, 56, 0.15) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Date Input                                            */
/* ══════════════════════════════════════════════════════ */
.stDateInput > div > div {
    border-radius: 12px !important;
    border-color: var(--lxt-border) !important;
    transition: all 0.3s ease !important;
}
.stDateInput > div > div:focus-within {
    border-color: var(--lxt-orange) !important;
    box-shadow: 0 0 0 3px rgba(254, 111, 56, 0.15) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Select Box / Number Input                             */
/* ══════════════════════════════════════════════════════ */
.stSelectbox > div > div,
.stNumberInput > div > div {
    border-radius: 12px !important;
    border-color: var(--lxt-border) !important;
    transition: all 0.3s ease !important;
}
.stSelectbox > div > div:focus-within,
.stNumberInput > div > div:focus-within {
    border-color: var(--lxt-orange) !important;
    box-shadow: 0 0 0 3px rgba(254, 111, 56, 0.15) !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Alerts / Info / Success / Warning / Error              */
/* ══════════════════════════════════════════════════════ */
.stAlert {
    border-radius: 12px !important;
    animation: fadeInUp 0.4s ease-out;
}
div[data-testid="stAlert"] {
    border-radius: 12px !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Custom Divider (gradient)                             */
/* ══════════════════════════════════════════════════════ */
hr {
    border: none !important;
    height: 1px !important;
    background: linear-gradient(90deg, transparent 0%, var(--lxt-orange) 20%, var(--lxt-teal) 80%, transparent 100%) !important;
    opacity: 0.4 !important;
    margin: 1.5rem 0 !important;
}

/* ══════════════════════════════════════════════════════ */
/*  Scrollbar                                             */
/* ══════════════════════════════════════════════════════ */
::-webkit-scrollbar {
    width: 6px;
    height: 6px;
}
::-webkit-scrollbar-track {
    background: var(--lxt-navy);
}
::-webkit-scrollbar-thumb {
    background: var(--lxt-teal);
    border-radius: 3px;
}
::-webkit-scrollbar-thumb:hover {
    background: var(--lxt-orange);
}

/* ══════════════════════════════════════════════════════ */
/*  Custom Classes                                        */
/* ══════════════════════════════════════════════════════ */
.lxt-login-card {
    background: rgba(13, 27, 42, 0.85);
    backdrop-filter: blur(20px);
    -webkit-backdrop-filter: blur(20px);
    border: 1px solid rgba(94, 139, 149, 0.25);
    border-radius: 20px;
    padding: 3rem 2.5rem;
    width: 440px;
    max-width: 90vw;
    box-shadow:
        0 20px 60px rgba(0, 0, 0, 0.4),
        0 0 40px rgba(254, 111, 56, 0.05),
        inset 0 1px 0 rgba(255, 255, 255, 0.05);
    animation: fadeInUp 0.8s cubic-bezier(0.16, 1, 0.3, 1);
}
.lxt-login-card:hover {
    border-color: rgba(254, 111, 56, 0.3);
    box-shadow:
        0 20px 60px rgba(0, 0, 0, 0.4),
        0 0 60px rgba(254, 111, 56, 0.08),
        inset 0 1px 0 rgba(255, 255, 255, 0.05);
}
.lxt-header {
    animation: fadeInUp 0.6s cubic-bezier(0.16, 1, 0.3, 1);
}
.lxt-section {
    animation: fadeInUp 0.5s ease-out;
}
.lxt-gradient-text {
    background: linear-gradient(135deg, var(--lxt-orange) 0%, var(--lxt-teal) 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
}
.lxt-badge {
    display: inline-block;
    background: rgba(254, 111, 56, 0.15);
    color: var(--lxt-orange);
    padding: 0.25rem 0.8rem;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}
.lxt-shimmer {
    background: linear-gradient(
        90deg,
        transparent 0%,
        rgba(254, 111, 56, 0.08) 50%,
        transparent 100%
    );
    background-size: 200% 100%;
    animation: shimmer 3s ease-in-out infinite;
}

/* ══════════════════════════════════════════════════════ */
/*  Chat Interface                                        */
/* ══════════════════════════════════════════════════════ */
.stChatMessage {
    animation: fadeInUp 0.4s ease-out;
    border-radius: 14px !important;
    margin-bottom: 0.5rem !important;
}
div[data-testid="stChatInput"] {
    border-radius: 50px !important;
}
div[data-testid="stChatInput"] textarea {
    border-radius: 50px !important;
    border-color: var(--lxt-border) !important;
    transition: all 0.3s ease !important;
    font-family: 'Inter', sans-serif !important;
}
div[data-testid="stChatInput"] textarea:focus {
    border-color: var(--lxt-orange) !important;
    box-shadow: 0 0 0 3px rgba(254, 111, 56, 0.15) !important;
}
.chat-header {
    background: var(--lxt-card-bg);
    backdrop-filter: blur(12px);
    border: 1px solid var(--lxt-border);
    border-radius: 16px;
    padding: 1.5rem;
    margin-bottom: 1rem;
    animation: fadeInUp 0.5s ease-out;
}
.chat-header:hover {
    border-color: rgba(254, 111, 56, 0.3);
    box-shadow: 0 4px 20px rgba(254, 111, 56, 0.08);
}

/* ══════════════════════════════════════════════════════ */
/*  Financial Dashboard                                   */
/* ══════════════════════════════════════════════════════ */
.kpi-card {
    background: var(--lxt-card-bg);
    backdrop-filter: blur(12px);
    border: 1px solid var(--lxt-border);
    border-radius: 16px;
    padding: 1.4rem 1.6rem;
    text-align: center;
    animation: fadeInUp 0.5s ease-out;
    transition: all 0.3s ease;
}
.kpi-card:hover {
    border-color: rgba(254, 111, 56, 0.3);
    box-shadow: 0 8px 30px rgba(254, 111, 56, 0.1);
    transform: translateY(-2px);
}
.kpi-label {
    font-size: 0.75rem;
    color: #8899A6;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    font-weight: 600;
    margin-bottom: 0.4rem;
}
.kpi-value {
    font-size: clamp(0.95rem, 1.4vw, 1.4rem);
    font-weight: 800;
    letter-spacing: -0.5px;
    margin: 0;
    line-height: 1.2;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.kpi-value.positive { color: #00C9A7; }
.kpi-value.negative { color: #FF6B6B; }
.kpi-value.neutral  { color: #F0F2F6; }
.kpi-value.accent   { color: #FE6F38; }
.kpi-sub {
    font-size: 0.7rem;
    color: #5A6C7D;
    margin-top: 0.3rem;
}
.dashboard-section-title {
    font-size: 1.1rem;
    font-weight: 700;
    color: #F0F2F6;
    letter-spacing: -0.3px;
    margin: 1.2rem 0 0.6rem 0;
    display: flex;
    align-items: center;
    gap: 0.5rem;
}
</style>

"""
st.markdown(_LXT_CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────
QB_TOKEN_URL = "https://oauth.platform.intuit.com/oauth2/v1/tokens/bearer"
QB_BASE_URL = "https://quickbooks.api.intuit.com"

# Path to the Streamlit secrets file (for auto-saving refresh tokens)
SECRETS_PATH = Path(__file__).parent / ".streamlit" / "secrets.toml"


def get_secret(key: str, default: str = "") -> str:
    """Read a secret from Streamlit secrets (secrets.toml) or environment variables.

    Works on all platforms:
      - Local / Streamlit Cloud: reads from .streamlit/secrets.toml
      - Render / Railway: reads from environment variables
    """
    try:
        value = st.secrets.get(key, None)
        if value is not None:
            return value
    except Exception:
        pass
    return os.environ.get(key, default)


def _load_companies() -> dict:
    """Load company credentials from secrets.toml or flat environment variables.

    On Streamlit Cloud / local: reads nested [companies.xxx] from secrets.toml.
    On Render / Railway: reads flat env vars (e.g. LXT_EGYPT_REALM_ID).
    """
    # Try Streamlit secrets first (nested TOML)
    try:
        companies = st.secrets.get("companies", None)
        if companies is not None:
            return dict(companies)
    except Exception:
        pass

    # Fall back to flat environment variables
    _COMPANY_ENV_KEYS = [
        ("lxt_egypt",           "LXT_EGYPT"),
        ("lxt_canada",          "LXT_CANADA"),
        ("lxt_australia",       "LXT_AUSTRALIA"),
        ("lxt_romania",         "LXT_ROMANIA"),
        ("lxt_india",           "LXT_INDIA"),
        ("lxt_germany",         "LXT_GERMANY"),
        ("lxt_uk",              "LXT_UK"),
        ("lxt_usa",             "LXT_USA"),
        ("lxt_clickworker_usa", "LXT_CLICKWORKER_USA"),
    ]

    companies = {}
    for company_key, env_prefix in _COMPANY_ENV_KEYS:
        label = os.environ.get(f"{env_prefix}_LABEL", "")
        realm_id = os.environ.get(f"{env_prefix}_REALM_ID", "")
        refresh_token = os.environ.get(f"{env_prefix}_REFRESH_TOKEN", "")
        if realm_id:
            companies[company_key] = {
                "label": label,
                "realm_id": realm_id,
                "refresh_token": refresh_token,
            }

    if companies:
        return companies

    logger.error("No company credentials found in secrets or environment.")
    return {}

# Company label → local currency (ISO codes)
COMPANY_CURRENCY = {
    "LXT Egypt": "EGP",
    "LXT Canada": "CAD",
    "LXT Australia": "AUD",
    "LXT Romania": "RON",
    "LXT India": "INR",
    "CW GmbH": "EUR",
    "LXT UK": "GBP",
    "LXT USA": "USD",
    "CW Inc": "USD",
}

# Currencies that need forex rate input (USD is always 1.0)
FOREX_CURRENCIES = ["EGP", "CAD", "AUD", "RON", "INR", "EUR", "GBP"]

OUTPUT_COLUMNS = [
    "Distribution account",
    "Account Number",
    "Transaction date",
    "Reporting Month",
    "Memo/Description",
    "Name",
    "Transaction id",
    "Customer full name",
    "Supplier",
    "Number",
    "Balance",
    "Debit",
    "Credit",
    "Class full name",
    "CostCenter",
    "SubClass Name",
    "Company Country",
    "Mapping",
    "Item",
    "Statement",
    "Transaction Value in Original Currency",
    "Currency",
    "Forex Rate",
    "Amount in USD (Reporting Currency)",
]

QB_COLUMN_MAP = {
    # Internal API keys
    "account_name": "Distribution account",
    "tx_date": "Transaction date",
    "memo": "Memo/Description",
    "name": "Name",
    "txn_type": "Transaction id",
    "cust_name": "Customer full name",
    "vend_name": "Supplier",
    "doc_num": "Number",
    "subt_nat_amount": "Balance",
    "subt_nat_home_amount": "Balance",
    "debt_amt": "Debit",
    "credit_amt": "Credit",
    "klass_name": "Class full name",

    # Display-name ColTitles returned by the API
    "Account": "Distribution account",
    "Distribution Account": "Distribution account",
    "Transaction Type": "Transaction id",
    "Trans #": "Transaction id",
    "No.": "Number",
    "Num": "Number",
    "Customer": "Customer full name",
    "Vendor": "Supplier",
    "Memo/Description": "Memo/Description",
    "Date": "Transaction date",
    "Class": "Class full name",

    "Amount": "Balance",
    "Debit": "Debit",
    "Credit": "Credit",
    # Home-currency debit/credit (from debt_home_amt / credit_home_amt)
    "Home Debit": "Debit",
    "Home Credit": "Credit",
    "debt_home_amt": "Debit",
    "credit_home_amt": "Credit",
    # For multi-currency companies the API labels transaction-currency
    # amounts as "Foreign Debit"/"Foreign Credit". We intentionally do NOT
    # map them to Debit/Credit — we want home-currency amounts instead.
    # "Foreign Debit" and "Foreign Credit" are kept as-is and ignored.
}

QB_REPORT_COLUMNS = "account_name,tx_date,memo,name,txn_type,cust_name,vend_name,doc_num,subt_nat_amount,subt_nat_home_amount,debt_amt,credit_amt,debt_home_amt,credit_home_amt,klass_name"


# ═══════════════════════════════════════════════════════════════
# Authentication UI
# ═══════════════════════════════════════════════════════════════
def check_password() -> bool:
    """Show a login form and return True if authenticated."""
    if st.session_state.get("authenticated"):
        # ── Session timeout (30 min inactivity) ──
        max_idle = 1800  # seconds
        now = time.time()
        last_activity = st.session_state.get("last_activity", now)
        if now - last_activity > max_idle:
            st.session_state.clear()
            st.warning("⏱️ Session expired due to inactivity. Please log in again.")
            st.rerun()
        st.session_state["last_activity"] = now
        return True

    # Background pattern
    st.markdown(
        """
        <style>
            .login-bg {
                position: fixed;
                top: 0; left: 0; right: 0; bottom: 0;
                z-index: -1;
                background:
                    radial-gradient(ellipse at 20% 50%, rgba(254,111,56,0.06) 0%, transparent 50%),
                    radial-gradient(ellipse at 80% 20%, rgba(94,139,149,0.06) 0%, transparent 50%),
                    radial-gradient(ellipse at 50% 80%, rgba(254,111,56,0.04) 0%, transparent 50%);
            }
        </style>
        <div class="login-bg"></div>
        """,
        unsafe_allow_html=True,
    )

    logo_html = ""
    if _logo_b64:
        logo_html = f'<img src="data:image/png;base64,{_logo_b64}" style="height:48px; margin-bottom:1.5rem;" alt="LXT Logo">'

    st.markdown(
        f"""
        <div style="display:flex; justify-content:center; align-items:center; min-height:70vh;">
            <div class="lxt-login-card">
                <div style="text-align:center;">
                    {logo_html}
                    <div style="margin-top:0.8rem; margin-bottom:1.2rem;"><span class="lxt-badge">Financial Reports</span></div>
                    <h2 style="
                        margin: 0 0 0.3rem 0;
                        font-size: 1.6rem;
                        font-weight: 700;
                        color: #F0F2F6;
                        letter-spacing: -0.5px;
                    ">Welcome Back</h2>
                    <p style="
                        color: #8899A6;
                        font-size: 0.9rem;
                        margin-bottom: 2rem;
                        font-weight: 400;
                    ">Enter your credentials to access the dashboard</p>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1.2, 1, 1.2])
    with col2:
        username = st.text_input(
            "Email",
            placeholder="Enter your email…",
            label_visibility="collapsed",
        )
        password = st.text_input(
            "Password",
            type="password",
            placeholder="Enter your password…",
            label_visibility="collapsed",
        )
        # ── Brute-force protection ──
        if "login_attempts" not in st.session_state:
            st.session_state["login_attempts"] = 0
            st.session_state["lockout_until"] = 0.0

        now = time.time()
        if now < st.session_state["lockout_until"]:
            remaining = int(st.session_state["lockout_until"] - now)
            mins, secs = divmod(remaining, 60)
            st.error(
                f"🔒 Too many failed attempts. "
                f"Please try again in **{mins}m {secs}s**."
            )
            return False

        login_clicked = st.button("Login", width="stretch", type="primary")

        if login_clicked:
            if not username.strip():
                st.error("❌ Please enter your email.")
                return False

            # ── Validate credentials ──
            username_ok = (
                username.strip().lower() == get_secret("APP_USERNAME").lower()
            )
            password_ok = False
            if username_ok:
                stored_hash = get_secret("APP_PASSWORD_HASH")
                try:
                    password_ok = bcrypt.checkpw(
                        password.encode("utf-8"),
                        stored_hash.encode("utf-8"),
                    )
                except (ValueError, TypeError) as exc:
                    st.error(
                        "⚠️ Password hash configuration error. "
                        "Please verify that `APP_PASSWORD_HASH` in Streamlit secrets "
                        "is a valid bcrypt hash (starts with `$2b$`)."
                    )
                    return False

            if username_ok and password_ok:
                # Successful login — reset attempts
                st.session_state["login_attempts"] = 0
                st.session_state["lockout_until"] = 0.0
                st.session_state["authenticated"] = True
                st.session_state["username"] = html.escape(username.strip())
                st.rerun()
            else:
                st.session_state["login_attempts"] += 1
                attempts_left = 5 - st.session_state["login_attempts"]
                if st.session_state["login_attempts"] >= 5:
                    st.session_state["lockout_until"] = time.time() + 900  # 15 min
                    st.error(
                        "🔒 Too many failed attempts. "
                        "Account locked for **15 minutes**."
                    )
                elif attempts_left > 0:
                    st.error(
                        f"❌ Incorrect email or password. "
                        f"**{attempts_left}** attempt(s) remaining."
                    )
                else:
                    st.error("❌ Incorrect email or password.")

    return False


# ═══════════════════════════════════════════════════════════════
# QuickBooks API Functions
# ═══════════════════════════════════════════════════════════════
def refresh_access_token(
    client_id: str, client_secret: str, refresh_token: str
) -> dict:
    """Exchange a refresh token for a fresh access token."""
    credentials = f"{client_id}:{client_secret}"
    auth_header = base64.b64encode(credentials.encode()).decode()

    resp = requests.post(
        QB_TOKEN_URL,
        headers={
            "Accept": "application/json",
            "Authorization": f"Basic {auth_header}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        data={"grant_type": "refresh_token", "refresh_token": refresh_token},
        timeout=30,
    )

    if resp.status_code != 200:
        raise RuntimeError(
            f"Token refresh failed (HTTP {resp.status_code}): {resp.text}"
        )

    data = resp.json()
    return {
        "access_token": data["access_token"],
        "refresh_token": data.get("refresh_token", refresh_token),
    }

# ═══════════════════════════════════════════════════════════════
# GitHub Gist — Persistent Token Storage
# ═══════════════════════════════════════════════════════════════
GIST_FILENAME = "lxt_qb_tokens.json"
GIST_API = "https://api.github.com/gists"


def _get_github_headers() -> dict:
    """Return GitHub API headers using the token from secrets."""
    token = get_secret("GITHUB_TOKEN")
    return {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
    }


def _find_token_gist() -> str | None:
    """Find the existing private gist ID by filename, or return None."""
    try:
        resp = requests.get(GIST_API, headers=_get_github_headers(), timeout=15)
        if resp.status_code == 200:
            for gist in resp.json():
                if GIST_FILENAME in gist.get("files", {}):
                    return gist["id"]
    except Exception as e:
        logger.warning("Failed to find token Gist: %s", e)
    return None


def _load_gist_tokens() -> dict | None:
    """Load refresh tokens from the private gist. Returns dict {company_key: token}."""
    gist_id = _find_token_gist()
    if not gist_id:
        return None
    try:
        resp = requests.get(
            f"{GIST_API}/{gist_id}", headers=_get_github_headers(), timeout=15
        )
        if resp.status_code == 200:
            import json
            content = resp.json()["files"][GIST_FILENAME]["content"]
            return json.loads(content)
    except Exception as e:
        logger.warning("Failed to load tokens from Gist: %s", e)
    return None


def _save_gist_tokens(tokens: dict) -> None:
    """Create or update the private gist with current tokens."""
    import json
    payload = {
        "description": "LXT QuickBooks refresh tokens (auto-managed)",
        "files": {
            GIST_FILENAME: {"content": json.dumps(tokens, indent=2)}
        },
    }
    try:
        gist_id = _find_token_gist()
        headers = _get_github_headers()
        if gist_id:
            requests.patch(
                f"{GIST_API}/{gist_id}",
                headers=headers,
                json=payload,
                timeout=15,
            )
        else:
            payload["public"] = False
            requests.post(
                GIST_API,
                headers=headers,
                json=payload,
                timeout=15,
            )
    except Exception as e:
        logger.error("⚠️ CRITICAL: Failed to save refresh tokens to Gist: %s", e)
        st.warning(
            f"⚠️ **Token save failed** — new refresh tokens could not be saved to Gist. "
            f"Error: {e}. If this persists, you may lose access to QuickBooks."
        )


def _save_refresh_token(old_token: str, new_token: str) -> None:
    """
    Replace the old refresh token with the new one in secrets.toml (local).
    This ensures the next run uses the latest single-use token.
    """
    if old_token == new_token or not SECRETS_PATH.exists():
        return

    try:
        content = SECRETS_PATH.read_text()
        if old_token in content:
            content = content.replace(old_token, new_token)
            SECRETS_PATH.write_text(content)
    except Exception as e:
        logger.warning("Failed to update local secrets.toml: %s", e)


def fetch_general_ledger(
    access_token: str, realm_id: str, start_date: str, end_date: str
) -> dict:
    """GET the General Ledger report JSON from QBO."""
    url = f"{QB_BASE_URL}/v3/company/{realm_id}/reports/GeneralLedger"
    resp = requests.get(
        url,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/json",
        },
        params={
            "start_date": start_date,
            "end_date": end_date,
            "columns": QB_REPORT_COLUMNS,
            "accounting_method": "Accrual",
        },
        timeout=120,
    )

    if resp.status_code != 200:
        raise RuntimeError(
            f"API request failed (HTTP {resp.status_code}): {resp.text}"
        )

    return resp.json()


# ═══════════════════════════════════════════════════════════════
# JSON Flattening (Recursive)
# ═══════════════════════════════════════════════════════════════
def flatten_report(report_json: dict) -> list[dict]:
    """Flatten the nested QBO report JSON into a list of row dicts."""
    columns_meta = report_json.get("Columns", {}).get("Column", [])
    col_keys = [c.get("ColTitle", "").strip() for c in columns_meta]

    rows = report_json.get("Rows", {}).get("Row", [])
    flat: list[dict] = []
    _walk(rows, col_keys, flat)
    return flat


def _walk(rows: list, col_keys: list[str], acc: list[dict]) -> None:
    """Recursively collect only type='Data' rows."""
    for row in rows:
        rtype = row.get("type", "Data")

        if rtype == "Data":
            cells = row.get("ColData", [])
            record = {
                col_keys[i] if i < len(col_keys) else f"col_{i}": c.get("value", "")
                for i, c in enumerate(cells)
            }
            acc.append(record)

        elif rtype == "Section":
            nested = row.get("Rows", {}).get("Row", [])
            if nested:
                _walk(nested, col_keys, acc)


# ═══════════════════════════════════════════════════════════════
# Pandas Transformation
# ═══════════════════════════════════════════════════════════════
def transform(raw_rows: list[dict], company_label: str) -> pd.DataFrame:
    """Rename columns, add Company Country, filter nulls."""
    if not raw_rows:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    df = pd.DataFrame(raw_rows)

    # Rename columns using the mapping.
    # For multi-currency companies the API returns "Foreign Debit" / "Foreign Credit"
    # and "Amount" (from subt_nat_home_amount) — all in the company's home currency.
    # The QB_COLUMN_MAP handles all variants (Foreign/Nat/plain) → Debit/Credit/Balance.
    rename_map = {
        k: v for k, v in QB_COLUMN_MAP.items() if k in df.columns and k != v
    }
    df = df.rename(columns=rename_map)

    # Enrich Transaction id
    if "Transaction id" in df.columns and "Number" in df.columns:
        df["Transaction id"] = (
            df["Transaction id"].astype(str).str.strip()
            + " #"
            + df["Number"].astype(str).str.strip()
        )

    # Ensure all columns exist
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df["Company Country"] = company_label
    df["Currency"] = COMPANY_CURRENCY.get(company_label, "")

    # Split "Class full name" on ":" into CostCenter and SubClass Name
    class_split = df["Class full name"].astype(str).str.split(":", n=1, expand=True)
    df["CostCenter"] = class_split[0].str.strip() if 0 in class_split.columns else ""
    df["SubClass Name"] = class_split[1].str.strip() if 1 in class_split.columns else ""

    # Reporting Month: month'year (e.g. "2'2026") derived from Transaction date
    td = pd.to_datetime(df["Transaction date"], format="mixed", errors="coerce")
    df["Reporting Month"] = td.dt.month.astype("Int64").astype(str) + "'" + td.dt.year.astype("Int64").astype(str)

    df = df[OUTPUT_COLUMNS]

    # Numeric conversion
    for c in ("Balance", "Debit", "Credit"):
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # Transaction Value in Original Currency:
    # If Debit has a value → -Debit, else if Credit has a value → Credit, else Balance
    import numpy as np
    df["Transaction Value in Original Currency"] = np.where(
        df["Debit"].notna(), -df["Debit"],
        np.where(df["Credit"].notna(), df["Credit"], df["Balance"])
    )

    # Filter empty Distribution account
    df["Distribution account"] = df["Distribution account"].astype(str).str.strip()
    df = df[
        (df["Distribution account"] != "")
        & (df["Distribution account"].str.lower() != "none")
        & (df["Distribution account"].str.lower() != "nan")
    ]
    df = df.dropna(subset=["Distribution account"]).reset_index(drop=True)

    return df


# ═══════════════════════════════════════════════════════════════
# Consol Mapping Sheet Lookup
# ═══════════════════════════════════════════════════════════════
def load_mapping(uploaded_file) -> pd.DataFrame:
    """Load the Consol Mapping sheet from an uploaded file and return a lookup DataFrame."""
    if uploaded_file is None:
        return pd.DataFrame(columns=["Account Number", "Mapping", "Item", "Statement"])

    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        mapping_df = pd.read_csv(uploaded_file)
    else:
        mapping_df = pd.read_excel(uploaded_file)

    # Normalise column names (strip whitespace)
    mapping_df.columns = mapping_df.columns.str.strip()
    # Ensure Account Number is string without float decimals (pandas reads as float64)
    mapping_df["Account Number"] = (
        mapping_df["Account Number"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    # Keep only the columns we need; drop duplicates so first match wins
    mapping_df = mapping_df[["Account Number", "Mapping", "Item", "Statement"]]
    mapping_df = mapping_df.drop_duplicates(subset="Account Number", keep="first")
    return mapping_df


def apply_mapping(df: pd.DataFrame, mapping_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extract the leading account number code from 'Distribution account'
    (e.g. '110205' from '110205 WISE RON') and merge with the Consol
    Mapping sheet to add Mapping, Item, and Statement columns.
    """
    if mapping_df.empty:
        return df

    # Extract leading numeric code from Distribution account
    df["_account_code"] = (
        df["Distribution account"]
        .astype(str)
        .str.extract(r"^(\d+)", expand=False)
        .str.strip()
    )

    # Merge on the account code
    df = df.merge(
        mapping_df,
        left_on="_account_code",
        right_on="Account Number",
        how="left",
        suffixes=("", "_map"),
    )

    # Keep Account Number from the mapping (the code used for lookup)
    if "Account Number_map" in df.columns:
        df["Account Number"] = df["Account Number_map"]
        df = df.drop(columns=["Account Number_map"])
    elif "Account Number" not in df.columns:
        df["Account Number"] = df["_account_code"]

    # If Mapping/Item/Statement columns already existed (as empty placeholders),
    # overwrite them with the merged values
    for col in ("Mapping", "Item", "Statement"):
        if f"{col}_map" in df.columns:
            df[col] = df[f"{col}_map"]
            df = df.drop(columns=[f"{col}_map"])

    # Clean up helper columns
    df = df.drop(columns=["_account_code"], errors="ignore")

    return df


# ═══════════════════════════════════════════════════════════════
# Forex Rate File Parser
# ═══════════════════════════════════════════════════════════════
def parse_forex_rate_file(uploaded_file) -> dict[str, dict[str, dict[str, float]]]:
    """
    Parse an uploaded Exchange Rate file (Excel or CSV) and return a
    nested lookup:
        { currency: { month_key: { "closing": rate, "average": rate } } }

    Expected columns:
      A  Currency
      B  End of Month       (date — used to derive month'year key)
      F  ClosingRate2       (closing rate vs USD)
      G  AverageRate2       (average rate vs USD)
    """
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    # Normalise column names
    df.columns = df.columns.str.strip()

    # Validate required columns exist
    required = {"Currency", "End of Month", "ClosingRate2", "AverageRate2"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Uploaded file is missing columns: {', '.join(missing)}")

    # Parse the End of Month date and build a month key like "1'2026"
    df["_eom"] = pd.to_datetime(df["End of Month"], format="mixed", errors="coerce")
    df["_month_key"] = (
        df["_eom"].dt.month.astype("Int64").astype(str)
        + "'"
        + df["_eom"].dt.year.astype("Int64").astype(str)
    )

    df["ClosingRate2"] = pd.to_numeric(df["ClosingRate2"], errors="coerce").fillna(1.0)
    df["AverageRate2"] = pd.to_numeric(df["AverageRate2"], errors="coerce").fillna(1.0)

    # Build lookup dict
    forex: dict[str, dict[str, dict[str, float]]] = {}
    for _, row in df.iterrows():
        ccy = str(row["Currency"]).strip().upper()
        mk = str(row["_month_key"]).strip()
        if not ccy or not mk or mk == "<NA>'<NA>":
            continue
        forex.setdefault(ccy, {})[mk] = {
            "closing": float(row["ClosingRate2"]),
            "average": float(row["AverageRate2"]),
        }

    # Ensure USD always resolves to 1.0
    if "USD" not in forex:
        forex["USD"] = {}

    return forex


# ═══════════════════════════════════════════════════════════════
# Excel Export (in-memory)
# ═══════════════════════════════════════════════════════════════
def to_excel_bytes(df: pd.DataFrame) -> bytes:
    """Write DataFrame to an in-memory Excel file and return bytes."""
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# Pivot P&L Report
# ═══════════════════════════════════════════════════════════════
def _month_key(year: int, month: int) -> str:
    """Return the Reporting Month key matching the format in the data, e.g. '1\'2026'."""
    return f"{month}'{year}"


def _prev_months(year: int, month: int, count: int) -> list[tuple[int, int]]:
    """Return a list of (year, month) tuples going back `count` months inclusive."""
    result = []
    for _ in range(count):
        result.append((year, month))
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    return result


def _classify_statements(master_df: pd.DataFrame) -> dict[str, str]:
    """
    Classify each Statement into 'revenue', 'cops', 'expenses', or 'other'
    based on the leading digit of its Account Numbers.
      4xxxxx → revenue,  5xxxxx → cops,  6xxxxx → expenses
    """
    classification: dict[str, str] = {}
    for stmt in master_df["Statement"].dropna().unique():
        stmt_str = str(stmt).strip()
        if not stmt_str or stmt_str.lower() == "nan":
            continue
        stmt_df = master_df[master_df["Statement"] == stmt]
        acct_nums = stmt_df["Account Number"].dropna().astype(str).str.strip()
        acct_nums = acct_nums[acct_nums != ""]
        if acct_nums.empty:
            classification[stmt_str] = "other"
            continue
        # Use the most common leading digit
        leading = acct_nums.str[0]
        mode_digit = leading.mode().iloc[0] if len(leading) > 0 else "0"
        if mode_digit == "4":
            classification[stmt_str] = "revenue"
        elif mode_digit == "5":
            classification[stmt_str] = "cops"
        elif mode_digit == "6":
            classification[stmt_str] = "expenses"
        else:
            classification[stmt_str] = "other"
    return classification


def _build_statement_rows(master_df: pd.DataFrame, stmt) -> list[dict]:
    """Build group header + detail lines + total row for a single Statement."""
    stmt_str = str(stmt).strip()
    stmt_df = master_df[master_df["Statement"] == stmt]
    rows: list[dict] = []

    # Statement group header
    rows.append({
        "Code": "",
        "Description": stmt_str,
        "_style": "group",
        "_statement": stmt,
        "_mapping": None,
    })

    # Mapping detail lines
    mappings = stmt_df["Mapping"].dropna().unique()
    for mapping in mappings:
        mapping_str = str(mapping).strip()
        if not mapping_str or mapping_str.lower() == "nan":
            continue

        # Find the most common Account Number for this mapping
        map_df = stmt_df[stmt_df["Mapping"] == mapping]
        acct_nums = map_df["Account Number"].dropna().astype(str).str.strip()
        acct_nums = acct_nums[acct_nums != ""]
        code = acct_nums.mode().iloc[0] if len(acct_nums) > 0 else ""

        rows.append({
            "Code": code,
            "Description": f"  {mapping_str}",
            "_style": "detail",
            "_statement": stmt,
            "_mapping": mapping,
        })

    # Statement total row
    rows.append({
        "Code": "",
        "Description": f"Total {stmt_str}",
        "_style": "total",
        "_statement": stmt,
        "_mapping": "__TOTAL__",
    })

    return rows


def _build_row_index(master_df: pd.DataFrame) -> list[dict]:
    """
    Build the shared row hierarchy with calculated rows.

    Order: Revenue stmts → COPS stmts → Gross Profit → GP% →
           Expense stmts → Total Expenses → Other stmts
    """
    classification = _classify_statements(master_df)
    statements = master_df["Statement"].dropna().unique()

    # Group statements by category
    revenue_stmts = [s for s in statements if classification.get(str(s).strip()) == "revenue"]
    cops_stmts = [s for s in statements if classification.get(str(s).strip()) == "cops"]
    expense_stmts = [s for s in statements if classification.get(str(s).strip()) == "expenses"]
    other_stmts = [s for s in statements if classification.get(str(s).strip()) == "other"]

    rows: list[dict] = []

    # ── Revenue statements ──
    for stmt in revenue_stmts:
        rows.extend(_build_statement_rows(master_df, stmt))

    # ── COPS statements ──
    for stmt in cops_stmts:
        rows.extend(_build_statement_rows(master_df, stmt))

    # ── Gross Profit (Revenue - COPS) ──
    rows.append({
        "Code": "4XXXXX - 5XXXXXX",
        "Description": "Gross Profit",
        "_style": "calculated",
        "_statement": "__CALCULATED__",
        "_mapping": "__GROSS_PROFIT__",
    })

    # ── GP% (Gross Profit / Revenue) ──
    rows.append({
        "Code": "(4XXXXX - 5XXXXXX) / 4XXXXXX",
        "Description": "Gross Profit %",
        "_style": "calculated",
        "_statement": "__CALCULATED__",
        "_mapping": "__GP_PCT__",
    })

    # ── Expense statements ──
    for stmt in expense_stmts:
        rows.extend(_build_statement_rows(master_df, stmt))

    # ── Total Expenses ──
    rows.append({
        "Code": "6XXXXXX",
        "Description": "Total Expenses",
        "_style": "calculated",
        "_statement": "__CALCULATED__",
        "_mapping": "__TOTAL_EXPENSES__",
    })

    # ── Other statements (non-operating, etc.) ──
    for stmt in other_stmts:
        rows.extend(_build_statement_rows(master_df, stmt))

    return rows


def _compute_section_values(
    df: pd.DataFrame,
    row_index: list[dict],
    month_keys: list[str],
    col_prefix: str,
    month_labels: list[str],
    classification: dict[str, str],
) -> tuple[list[str], dict[int, dict[str, float]]]:
    """
    Compute values for one section (Consolidated / Entity / CostCenter).

    Returns:
      - col_names: list of column names [prefix M1, prefix M2, prefix M3, prefix Var]
      - values: dict mapping row_index position → {col_name: value}
    """
    col_names = [f"{col_prefix} {lbl}" for lbl in month_labels] + [f"{col_prefix} Variance"]
    values: dict[int, dict[str, float]] = {}

    # Track totals per statement for total rows
    stmt_totals: dict[str, dict[str, float]] = {}

    for idx, row in enumerate(row_index):
        style = row["_style"]
        stmt = row["_statement"]
        mapping = row["_mapping"]

        if style == "group":
            # Group header — no values
            values[idx] = {c: None for c in col_names}
            continue

        if style == "detail" and mapping is not None:
            row_vals = {}
            for i, mk in enumerate(month_keys):
                cn = col_names[i]
                mask = (
                    (df["Statement"] == stmt)
                    & (df["Mapping"] == mapping)
                    & (df["Reporting Month"] == mk)
                )
                val = df.loc[mask, "Amount in USD (Reporting Currency)"].sum()
                row_vals[cn] = round(val, 2)

                # Accumulate statement totals
                total_key = str(stmt)
                if total_key not in stmt_totals:
                    stmt_totals[total_key] = {c: 0.0 for c in col_names[:-1]}
                stmt_totals[total_key][cn] += val

            # Variance = latest month - previous month
            row_vals[col_names[-1]] = round(row_vals[col_names[0]] - row_vals[col_names[1]], 2)
            values[idx] = row_vals

        elif style == "total" and mapping == "__TOTAL__":
            total_key = str(stmt)
            totals = stmt_totals.get(total_key, {c: 0.0 for c in col_names[:-1]})
            row_vals = {c: round(totals.get(c, 0.0), 2) for c in col_names[:-1]}
            row_vals[col_names[-1]] = round(row_vals[col_names[0]] - row_vals[col_names[1]], 2)
            values[idx] = row_vals

        elif style == "calculated":
            # ── Aggregate by category ──
            def _cat_sum(category: str) -> dict[str, float]:
                """Sum all statement totals belonging to a category."""
                result = {c: 0.0 for c in col_names[:-1]}
                for s_key, s_vals in stmt_totals.items():
                    if classification.get(s_key) == category:
                        for c in col_names[:-1]:
                            result[c] += s_vals.get(c, 0.0)
                return result

            if mapping == "__GROSS_PROFIT__":
                rev = _cat_sum("revenue")
                cops = _cat_sum("cops")
                row_vals = {c: round(rev[c] - cops[c], 2) for c in col_names[:-1]}
                row_vals[col_names[-1]] = round(
                    row_vals[col_names[0]] - row_vals[col_names[1]], 2
                )
                values[idx] = row_vals

            elif mapping == "__GP_PCT__":
                rev = _cat_sum("revenue")
                cops = _cat_sum("cops")
                row_vals = {}
                for c in col_names[:-1]:
                    gp = rev[c] - cops[c]
                    row_vals[c] = f"{(gp / rev[c] * 100):.1f}%" if rev[c] != 0 else "0.0%"
                # Variance for GP% — difference in percentage points
                gp_latest = (rev[col_names[0]] - cops[col_names[0]])
                gp_prev = (rev[col_names[1]] - cops[col_names[1]])
                pct_latest = (gp_latest / rev[col_names[0]] * 100) if rev[col_names[0]] != 0 else 0
                pct_prev = (gp_prev / rev[col_names[1]] * 100) if rev[col_names[1]] != 0 else 0
                row_vals[col_names[-1]] = f"{(pct_latest - pct_prev):.1f}pp"
                values[idx] = row_vals

            elif mapping == "__TOTAL_EXPENSES__":
                exp = _cat_sum("expenses")
                row_vals = {c: round(exp[c], 2) for c in col_names[:-1]}
                row_vals[col_names[-1]] = round(
                    row_vals[col_names[0]] - row_vals[col_names[1]], 2
                )
                values[idx] = row_vals

    return col_names, values


def build_pivot_report(
    master_df: pd.DataFrame,
    selected_year: int,
    selected_month: int,
) -> tuple[pd.DataFrame, list[dict], list[str], list[tuple[str, list[str]]]]:
    """
    Build the full horizontal pivot P&L report.

    Returns:
      - display_df: DataFrame for Streamlit preview
      - raw_rows: list of row dicts with '_style' metadata
      - all_columns: ordered list of all column names
      - section_groups: list of (section_label, [col_names]) for Excel header grouping
    """
    import calendar

    # Determine 3 consecutive months (latest first)
    months = _prev_months(selected_year, selected_month, 3)
    month_keys = [_month_key(y, m) for y, m in months]
    month_labels = [calendar.month_abbr[m] + f" {y}" for y, m in months]

    pl_df = master_df.copy()

    # Keep only P&L items (exclude B.S / Balance Sheet rows)
    pl_df = pl_df[pl_df["Item"].astype(str).str.strip().str.upper() == "P&L"]

    # Build shared row hierarchy
    row_index = _build_row_index(pl_df)
    classification = _classify_statements(pl_df)

    # Collect all section column groups
    all_section_cols: list[str] = []
    section_groups: list[tuple[str, list[str]]] = []

    # Helper to add a section
    def add_section(data_df, prefix):
        col_names, vals = _compute_section_values(
            data_df, row_index, month_keys, prefix, month_labels, classification
        )
        all_section_cols.extend(col_names)
        section_groups.append((prefix, col_names))
        return vals

    # ── Section 1: Consolidated ──
    all_values = [add_section(pl_df, "Consolidated")]

    # ── Section 2: Per Legal Entity ──
    entities = sorted(pl_df["Company Country"].dropna().unique())
    for entity in entities:
        entity_df = pl_df[pl_df["Company Country"] == entity]
        all_values.append(add_section(entity_df, entity))

    # ── Section 3: Per CostCenter ──
    cost_centers = pl_df["CostCenter"].dropna().astype(str).str.strip()
    cost_centers = sorted(cost_centers[cost_centers != ""].unique())
    for cc in cost_centers:
        cc_df = pl_df[pl_df["CostCenter"].astype(str).str.strip() == cc]
        all_values.append(add_section(cc_df, f"CC: {cc}"))

    # Build final rows
    all_columns = ["Code", "Description"] + all_section_cols
    raw_rows: list[dict] = []

    for idx, row in enumerate(row_index):
        out = {
            "Code": row["Code"],
            "Description": row["Description"],
            "_style": row["_style"],
        }
        for section_vals in all_values:
            if idx in section_vals:
                out.update(section_vals[idx])
            # Missing values default to empty for group rows
        raw_rows.append(out)

    # Build display DataFrame
    display_df = pd.DataFrame(raw_rows)
    display_cols = [c for c in all_columns if c in display_df.columns]
    display_df = display_df[display_cols]

    # Cast value columns to string so Arrow doesn't choke on
    # mixed numeric / percentage-string types (e.g. GP% rows).
    value_cols = [c for c in display_cols if c not in ("Code", "Description")]
    for vc in value_cols:
        display_df[vc] = display_df[vc].apply(
            lambda v: f"{v:,.2f}" if isinstance(v, (int, float)) and v == v else
                      (str(v) if v is not None else "")
        )

    return display_df, raw_rows, all_columns, section_groups


def pivot_to_excel_bytes(
    rows: list[dict],
    columns: list[str],
    section_groups: list[tuple[str, list[str]]],
    month_labels: list[str],
) -> bytes:
    """Write the horizontal pivot report to a styled Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot P&L Report"

    # ── Styles ──
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="2D2D44", end_color="2D2D44", fill_type="solid")
    group_font = Font(bold=True, size=10, color="1B3A5C")
    group_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    calc_font = Font(bold=True, size=11, color="1A1A00")
    calc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="3A3A5C", end_color="3A3A5C", fill_type="solid")
    section_header_fills = [
        PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid"),
        PatternFill(start_color="7D3C98", end_color="7D3C98", fill_type="solid"),
        PatternFill(start_color="1E8449", end_color="1E8449", fill_type="solid"),
        PatternFill(start_color="B9770E", end_color="B9770E", fill_type="solid"),
        PatternFill(start_color="922B21", end_color="922B21", fill_type="solid"),
    ]
    num_fmt = '#,##0.00'
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Row 1: Section group headers (merged) ──
    # Code + Description stay empty on row 1
    ws.cell(row=1, column=1, value="Code").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=2, value="Description").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = center_align

    col_offset = 3  # sections start at column C
    for sec_idx, (sec_label, sec_cols) in enumerate(section_groups):
        fill = section_header_fills[sec_idx % len(section_header_fills)]
        start_col = col_offset
        end_col = col_offset + len(sec_cols) - 1

        # Merge section header
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=sec_label)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = fill
        cell.alignment = center_align

        col_offset += len(sec_cols)

    # ── Row 2: Month sub-headers ──
    ws.cell(row=2, column=1, value="").fill = header_fill
    ws.cell(row=2, column=2, value="").fill = header_fill

    col_offset = 3
    for sec_idx, (sec_label, sec_cols) in enumerate(section_groups):
        for i, col_name in enumerate(sec_cols):
            cell = ws.cell(row=2, column=col_offset + i)
            # Extract the sub-label (remove the section prefix)
            parts = col_name.split(" ", 1)
            if len(parts) > 1:
                # Remove the prefix before the month label
                prefix = sec_label + " "
                sub_label = col_name[len(prefix):] if col_name.startswith(prefix) else col_name
            else:
                sub_label = col_name
            cell.value = sub_label
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        col_offset += len(sec_cols)

    # ── Data rows (starting row 3) ──
    for row_idx, row_data in enumerate(rows, start=3):
        style = row_data.get("_style", "detail")

        for col_idx, col_name in enumerate(columns, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if style == "group":
                cell.font = group_font
                cell.fill = group_fill
            elif style == "total":
                cell.font = total_font
                cell.fill = total_fill
            elif style == "calculated":
                cell.font = calc_font
                cell.fill = calc_fill
            elif style == "detail":
                cell.border = thin_border

            # Number formatting for value columns (col 3+)
            if col_idx > 2 and isinstance(val, (int, float)):
                cell.number_format = num_fmt
                cell.alignment = right_align

    # ── Column widths ──
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    for col_idx in range(3, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    # Freeze panes: freeze Code + Description columns and the 2 header rows
    ws.freeze_panes = "C3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# Main App
# ═══════════════════════════════════════════════════════════════
def main_app():
    """Render the main application after authentication."""

    # ── Sidebar ───────────────────────────────────────────────
    with st.sidebar:
        if _logo_b64:
            st.markdown(
                f"""
                <div style="text-align:center; padding:1.2rem 0 0.8rem 0; animation: fadeInUp 0.5s ease-out;">
                    <img src="data:image/png;base64,{_logo_b64}" style="height:36px;" alt="LXT">
                    <div style="
                        margin-top:0.6rem;
                        font-size:0.65rem;
                        font-weight:600;
                        letter-spacing:2px;
                        text-transform:uppercase;
                        color:#5E8B95;
                    ">FINANCIAL REPORTS</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
        st.divider()
        st.markdown(
            f"""
            <div style="animation: slideInLeft 0.4s ease-out;">
                <div style="
                    display:flex;
                    align-items:center;
                    gap:0.6rem;
                    padding:0.6rem 0.8rem;
                    background:rgba(94,139,149,0.08);
                    border-radius:12px;
                    margin-bottom:0.5rem;
                ">
                    <div style="
                        width:36px;
                        height:36px;
                        border-radius:10px;
                        background:linear-gradient(135deg, #FE6F38, #E5592A);
                        display:flex;
                        align-items:center;
                        justify-content:center;
                        font-size:1rem;
                        color:white;
                        font-weight:700;
                    ">A</div>
                    <div>
                        <div style="font-size:0.85rem; font-weight:600; color:#F0F2F6;">{st.session_state.get('username', 'Admin')}</div>
                        <div style="font-size:0.7rem; color:#8899A6;">Administrator</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.divider()
        if st.button("🚪 Logout", width="stretch"):
            st.session_state.clear()
            st.rerun()

    # ── Header ────────────────────────────────────────────────
    header_logo = ""
    if _logo_b64:
        header_logo = f'<img src="data:image/png;base64,{_logo_b64}" style="height:42px;" alt="LXT">'

    st.markdown(
        f"""
        <div class="lxt-header" style="margin-bottom:0.5rem;">
            <div style="display:flex; align-items:center; gap:1rem; margin-bottom:0.5rem;">
                {header_logo}
                <div>
                    <h1 style="
                        margin:0;
                        font-size:2rem;
                        font-weight:800;
                        letter-spacing:-1px;
                        color:#F0F2F6;
                    ">Financial Consolidated Report</h1>
                </div>
            </div>
            <p style="
                color:#8899A6;
                font-size:0.95rem;
                margin:0.5rem 0 0 0;
                max-width:700px;
                line-height:1.5;
            ">Extract General Ledger data from <strong style="color:#FE6F38;">9 QuickBooks companies</strong>
               and download a single consolidated Excel report.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.divider()

    # ── Date Inputs ───────────────────────────────────────────
    today = date.today()
    first_of_this_month = today.replace(day=1)

    if today.month == 1:
        default_end = first_of_this_month - timedelta(days=1)
        default_start = default_end.replace(month=1, day=1)
    else:
        default_end = first_of_this_month - timedelta(days=1)
        default_start = today.replace(month=1, day=1)

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        start_date = st.date_input("📅 Start Date", value=default_start)
    with col2:
        end_date = st.date_input("📅 End Date", value=default_end)

    if start_date > end_date:
        st.error("Start date cannot be after end date.")
        return

    st.divider()

    # ── Consol Mapping File Upload ─────────────────────────────
    with st.expander("📑 Consol Mapping File", expanded=True):
        st.caption(
            "Upload the **Consol Mapping Sheet** file (Excel or CSV). "
            "The file should contain columns: **Account Number**, "
            "**Mapping**, **Item**, **Statement**."
        )
        mapping_file = st.file_uploader(
            "Upload Consol Mapping File",
            type=["xlsx", "xls", "csv"],
            key="mapping_file_upload",
            label_visibility="collapsed",
        )

        mapping_df = pd.DataFrame(columns=["Account Number", "Mapping", "Item", "Statement"])
        if mapping_file is not None:
            try:
                mapping_df = load_mapping(mapping_file)
                st.success(
                    f"✅ Loaded **{len(mapping_df)}** account mappings."
                )
            except Exception as exc:
                st.error(f"❌ Failed to parse mapping file: {exc}")
                mapping_df = pd.DataFrame(columns=["Account Number", "Mapping", "Item", "Statement"])
        else:
            st.info("ℹ️ No mapping file uploaded — Mapping, Item, and Statement columns will be empty.")

    st.divider()

    # ── Forex Rate File Upload ────────────────────────────────
    with st.expander("💱 Exchange Rate File", expanded=True):
        st.caption(
            "Upload the **Consolidated Exchange Rate** file (Excel or CSV). "
            "The file should contain columns: **Currency** (A), "
            "**End of Month** (B), **ClosingRate2** (F), **AverageRate2** (G). "
            "Rates will be matched automatically by currency and month."
        )
        forex_file = st.file_uploader(
            "Upload Exchange Rate File",
            type=["xlsx", "xls", "csv"],
            key="forex_file_upload",
            label_visibility="collapsed",
        )

        forex_rates: dict[str, dict[str, dict[str, float]]] = {}
        if forex_file is not None:
            try:
                forex_rates = parse_forex_rate_file(forex_file)
                # Show summary of loaded rates
                currencies_loaded = sorted(forex_rates.keys())
                months_count = max(len(v) for v in forex_rates.values()) if forex_rates else 0
                st.success(
                    f"✅ Loaded rates for **{len(currencies_loaded)}** currencies "
                    f"across **{months_count}** months: {', '.join(currencies_loaded)}"
                )
            except Exception as exc:
                st.error(f"❌ Failed to parse exchange rate file: {exc}")
                forex_rates = {}
        else:
            st.info("ℹ️ No exchange rate file uploaded — all forex rates will default to **1.0**.")

    st.divider()

    # ── Generate Button ─────────────────────────────────────
    generate = st.button(
        "🚀 Generate Report",
        type="primary",
        width="stretch",
    )

    if generate:
        _run_etl(
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d"),
            forex_rates,
            mapping_df,
        )

    # ── Show report results (persisted in session state) ──────
    if "report_data" in st.session_state and "report_name" in st.session_state:
        st.divider()

        col1, col2 = st.columns(2)
        col1.metric("Total Rows", f"{st.session_state['report_rows']:,}")
        col2.metric("File", st.session_state["report_name"])

        with st.expander("📋 Preview Data (first 100 rows)", expanded=True):
            st.dataframe(st.session_state["report_preview"], width="stretch")

        # Ensure filename is a plain string with .xlsx extension
        fname = str(st.session_state["report_name"])
        if not fname.endswith(".xlsx"):
            fname += ".xlsx"

        # Download button with Content-Disposition via Streamlit
        st.download_button(
            label="📥 Download Excel Report",
            data=st.session_state["report_data"],
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="download_report_btn",
        )

    # ── Pivot P&L Report Section (auto-generated) ──────────────
    if "master_df" in st.session_state:
        st.divider()
        st.subheader("📈 Pivot P&L Report")

        # Auto-detect latest month from the data
        if "pivot_data" not in st.session_state:
            import calendar as _cal_auto
            mdf = st.session_state["master_df"]
            reporting_months = mdf["Reporting Month"].dropna().unique().tolist()

            # Parse month keys like "1'2026" → (year, month)
            parsed_months = []
            for mk in reporting_months:
                try:
                    parts = str(mk).split("'")
                    m, y = int(parts[0]), int(parts[1])
                    parsed_months.append((y, m))
                except (ValueError, IndexError):
                    continue

            if len(parsed_months) < 3:
                st.warning(
                    f"⚠️ Only **{len(parsed_months)} month(s)** of data available "
                    f"({', '.join(str(mk) for mk in reporting_months)}). "
                    f"The Pivot P&L requires **3 consecutive months** for variance analysis. "
                    f"Some columns may show zero values."
                )

            if parsed_months:
                # Use the latest month
                latest = max(parsed_months, key=lambda t: (t[0], t[1]))
                pivot_year, pivot_month = latest

                with st.spinner("Building pivot report…"):
                    display_df, raw_rows, all_columns, section_groups = build_pivot_report(
                        mdf, pivot_year, pivot_month,
                    )
                    _m_tuples = _prev_months(pivot_year, pivot_month, 3)
                    _m_labels = [_cal_auto.month_abbr[m] + f" {y}" for y, m in _m_tuples]

                    pivot_xlsx = pivot_to_excel_bytes(
                        raw_rows, all_columns, section_groups, _m_labels
                    )
                    m_labels = [_cal_auto.month_abbr[m] + f"{y}" for y, m in _m_tuples]
                    pivot_fname = f"LXT_Pivot_PL_{m_labels[0]}_to_{m_labels[2]}.xlsx"

                    st.session_state["pivot_data"] = pivot_xlsx
                    st.session_state["pivot_name"] = pivot_fname
                    st.session_state["pivot_preview"] = display_df
                    st.session_state["pivot_rows"] = len(display_df)

        # Show persisted pivot results
        if "pivot_data" in st.session_state:
            col1, col2 = st.columns(2)
            col1.metric("Pivot Rows", f"{st.session_state['pivot_rows']:,}")
            col2.metric("File", st.session_state["pivot_name"])

            with st.expander("📋 Pivot Preview (first 200 rows)", expanded=True):
                st.dataframe(
                    st.session_state["pivot_preview"].head(200),
                    width="stretch",
                )

            pivot_fname = str(st.session_state["pivot_name"])
            if not pivot_fname.endswith(".xlsx"):
                pivot_fname += ".xlsx"

            st.download_button(
                label="📥 Download Pivot Report",
                data=st.session_state["pivot_data"],
                file_name=pivot_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                key="download_pivot_btn",
            )

    # ═══════════════════════════════════════════════════════════════
    # Financial Analytics Dashboard
    # ═══════════════════════════════════════════════════════════════
    if "master_df" in st.session_state:
        st.divider()
        st.subheader("📊 Financial Analytics Dashboard")
        _render_financial_dashboard(st.session_state["master_df"])


    # ═══════════════════════════════════════════════════════════════
    # AI Financial Chatbot
    # ═══════════════════════════════════════════════════════════════
    if "master_df" in st.session_state:
        st.divider()

        # ── Chat Header ──
        st.markdown(
            """
            <div class="chat-header">
                <div style="display:flex; align-items:center; gap:0.8rem;">
                    <div style="
                        width:44px; height:44px;
                        border-radius:12px;
                        background: linear-gradient(135deg, #FE6F38, #E5592A);
                        display:flex; align-items:center; justify-content:center;
                        font-size:1.4rem;
                        box-shadow: 0 4px 15px rgba(254, 111, 56, 0.3);
                    ">🤖</div>
                    <div>
                        <h3 style="margin:0; font-size:1.15rem; font-weight:700; color:#F0F2F6; letter-spacing:-0.3px;">
                            AI Financial Assistant
                        </h3>
                        <p style="margin:0; font-size:0.8rem; color:#8899A6;">
                            Ask questions about your consolidated data &amp; pivot report — powered by Google Gemini
                        </p>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        # ── Initialize chat history ──
        if "messages" not in st.session_state:
            st.session_state["messages"] = [
                {
                    "role": "assistant",
                    "content": (
                        "👋 Hello! I'm your AI Financial Assistant. I can analyze "
                        "your consolidated General Ledger and Pivot P&L data. \n\n"
                        "Here are some example questions:\n"
                        "- *What is the total revenue and expenses for each company?*\n"
                        "- *Break down LXT Egypt's debits and credits by month*\n"
                        "- *Which cost center has the highest spending in USD?*\n"
                        "- *Who are the top 10 suppliers by total amount?*\n"
                        "- *Compare the P&L performance across all entities*\n"
                        "- *What is the total Amount in USD by mapping category?*"
                    ),
                }
            ]

        # ── Display existing chat messages ──
        for msg in st.session_state["messages"]:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

        # ── Chat input ──
        if prompt := st.chat_input(
            "Ask a question about the financial data…"
        ):
            # Append and display user message
            st.session_state["messages"].append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)

            # ── Call Gemini directly ──
            with st.chat_message("assistant"):
                with st.spinner("Analyzing data…"):
                    try:
                        from google import genai
                        from google.genai import types

                        api_key = get_secret("GOOGLE_API_KEY")
                        if not api_key or api_key == "your_api_key_here":
                            response = (
                                "⚠️ **Google API Key not configured.** "
                                "Please add your `GOOGLE_API_KEY` to "
                                "`.streamlit/secrets.toml` to enable the AI assistant."
                            )
                        else:
                            # Build / cache the data context
                            if "financial_context" not in st.session_state:
                                pivot_df = st.session_state.get("pivot_preview")
                                forex = st.session_state.get("forex_rates", {})
                                mapping = st.session_state.get("mapping_df", pd.DataFrame())
                                st.session_state["financial_context"] = (
                                    _build_financial_context(
                                        st.session_state["master_df"],
                                        pivot_df,
                                        forex,
                                        mapping,
                                    )
                                )

                            # Detect company in the question → append filtered detail
                            extra_context = _detect_and_filter(
                                prompt, st.session_state["master_df"]
                            )

                            system_prompt = (
                                "You are a senior financial analyst assistant for LXT, a global company "
                                "with operations in Egypt, Canada, Australia, Romania, India, Germany (CW GmbH), "
                                "UK, and USA (including CW Inc).\n\n"
                                "You have access to the COMPLETE consolidated General Ledger data, "
                                "exchange rates, account mapping, and Pivot P&L report below. "
                                "Use ONLY this data to answer questions — never estimate, guess, or use external sources.\n\n"
                                "STRICT RULES:\n"
                                "1. EXCHANGE RATES: You must ONLY use the exchange rates provided in the "
                                "   data below. NEVER use any other exchange rate from your training data "
                                "   or external knowledge. If a rate is not in the provided data, say so.\n"
                                "2. NEGATIVE EXPENSES: In the General Ledger, expenses (COGS, SG&A, etc.) "
                                "   appear as NEGATIVE amounts. The negative sign is for accounting convention only. "
                                "   When calculating Gross Profit = Revenue - COGS, use the ABSOLUTE value of COGS. "
                                "   Example: Revenue = 100,000, COGS = -60,000 → Gross Profit = 100,000 - 60,000 = 40,000 "
                                "   (subtract the absolute value, do NOT add the negative).\n"
                                "3. P&L CALCULATIONS:\n"
                                "   - Gross Profit = Revenue - |COGS| (absolute value of COGS)\n"
                                "   - Gross Profit % = Gross Profit / Revenue × 100\n"
                                "   - Net Income = Revenue - |COGS| - |Operating Expenses|\n"
                                "   - All P&L items use the AVERAGE exchange rate for the month\n"
                                "   - All Balance Sheet items use the CLOSING exchange rate for the month\n"
                                "4. FORMAT: Numbers with commas and 2 decimal places (e.g., $1,234,567.89 USD).\n"
                                "5. Default to USD as the reporting currency.\n"
                                "6. Present answers in a clean, professional format suitable for finance executives.\n"
                                "7. When comparing companies or periods, use markdown tables for clarity.\n"
                                "8. If the data doesn't contain enough detail to answer, say so explicitly.\n"
                                "9. Keep answers concise but complete.\n\n"
                                "FINANCIAL DATA:\n\n"
                                + st.session_state["financial_context"]
                                + extra_context
                            )

                            # Build multi-turn conversation (skip welcome message)
                            contents = []
                            for msg in st.session_state["messages"][1:]:
                                role = "model" if msg["role"] == "assistant" else "user"
                                contents.append(
                                    types.Content(
                                        role=role,
                                        parts=[types.Part(text=msg["content"])],
                                    )
                                )

                            client = genai.Client(api_key=api_key)
                            result = client.models.generate_content(
                                model="gemini-2.0-flash",
                                config=types.GenerateContentConfig(
                                    system_instruction=system_prompt,
                                    temperature=0.1,
                                ),
                                contents=contents,
                            )
                            response = result.text

                    except ImportError:
                        response = (
                            "⚠️ **Missing dependencies.** Please install:\n\n"
                            "```\npip install google-genai\n```"
                        )
                    except Exception as exc:
                        response = f"❌ **Error:** {exc}\n\nPlease try rephrasing your question."

                st.markdown(response)
                st.session_state["messages"].append(
                    {"role": "assistant", "content": response}
                )


# ═══════════════════════════════════════════════════════════════
# AI — Smart Company Detection & Filtering
# ═══════════════════════════════════════════════════════════════

# Keyword → actual "Company Country" value in the DataFrame.
# Multiple keywords can map to the same company.
_COMPANY_KEYWORDS: dict[str, str] = {
    "egypt": "LXT Egypt",
    "lxt egypt": "LXT Egypt",
    "canada": "LXT Canada",
    "lxt canada": "LXT Canada",
    "australia": "LXT Australia",
    "lxt australia": "LXT Australia",
    "romania": "LXT Romania",
    "lxt romania": "LXT Romania",
    "india": "LXT India",
    "lxt india": "LXT India",
    "germany": "CW GmbH",
    "gmbh": "CW GmbH",
    "cw gmbh": "CW GmbH",
    "uk": "LXT UK",
    "lxt uk": "LXT UK",
    "united kingdom": "LXT UK",
    "usa": "LXT USA",
    "lxt usa": "LXT USA",
    "cw inc": "CW Inc",
    "clickworker": "CW Inc",
}


def _detect_and_filter(prompt: str, master_df: pd.DataFrame) -> str:
    """
    Detect company references in the user's question.
    If exactly ONE company is mentioned, return a supplementary context
    string with that company's filtered raw data.
    Returns empty string if no single company is detected.
    """
    prompt_lower = prompt.lower()

    # Find all matching companies (deduplicated)
    matched: set[str] = set()
    # Sort by key length descending so "lxt egypt" matches before "egypt"
    for keyword in sorted(_COMPANY_KEYWORDS, key=len, reverse=True):
        if keyword in prompt_lower:
            matched.add(_COMPANY_KEYWORDS[keyword])

    # Only filter when exactly one company is detected
    if len(matched) != 1:
        return ""

    company = matched.pop()
    filtered = master_df[master_df["Company Country"] == company]

    if filtered.empty:
        return ""

    # Ensure numeric
    df = filtered.copy()
    for col in ["Debit", "Credit", "Balance", "Amount in USD (Reporting Currency)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Replace blanks
    for col in ["CostCenter", "Supplier", "Customer full name", "Mapping",
                "Distribution account", "Statement"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()
            df[col] = df[col].replace("", "Unassigned")

    num_cols = [
        c for c in ["Debit", "Credit", "Amount in USD (Reporting Currency)"]
        if c in df.columns
    ]

    ctx = []
    sep = "─" * 50

    ctx.append(f"\n{'═' * 60}")
    ctx.append(f"DETAILED DATA FOR: {company}")
    ctx.append("═" * 60)
    ctx.append(f"Transactions: {len(df):,}")

    # Per-month breakdown
    ctx.append(f"\n{sep}")
    ctx.append(f"{company} — BY MONTH")
    ctx.append(df.groupby("Reporting Month")[num_cols].sum().round(2).to_string())

    # Per-account breakdown
    ctx.append(f"\n{sep}")
    ctx.append(f"{company} — BY ACCOUNT")
    ctx.append(df.groupby("Distribution account")[num_cols].sum().round(2).to_string())

    # Per-cost-center
    if "CostCenter" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append(f"{company} — BY COST CENTER")
        ctx.append(df.groupby("CostCenter")[num_cols].sum().round(2).to_string())

    # Per-mapping
    if "Mapping" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append(f"{company} — BY MAPPING")
        ctx.append(
            df.groupby("Mapping")["Amount in USD (Reporting Currency)"]
            .sum().round(2).to_string()
        )

    # Month × Account
    ctx.append(f"\n{sep}")
    ctx.append(f"{company} — BY MONTH × ACCOUNT")
    ctx.append(
        df.groupby(["Reporting Month", "Distribution account"])[num_cols]
        .sum().round(2).to_string()
    )

    # Include raw data (capped to stay within token limits)
    MAX_COMPANY_ROWS = 500
    if len(df) <= MAX_COMPANY_ROWS:
        ctx.append(f"\n{sep}")
        ctx.append(f"{company} — ALL TRANSACTIONS (CSV)")
        ctx.append(df.to_csv(index=False))
    else:
        ctx.append(f"\n({company} has {len(df):,} rows — use the aggregated summaries above for totals.)")

    return "\n".join(ctx)


# ═══════════════════════════════════════════════════════════════
# AI Context Builder
# ═══════════════════════════════════════════════════════════════
def _build_financial_context(
    master_df: pd.DataFrame,
    pivot_df: pd.DataFrame | None = None,
    forex_rates: dict | None = None,
    mapping_df: pd.DataFrame | None = None,
) -> str:
    """
    Pre-aggregate financial data into a compact text summary.
    This is passed as system context to Gemini so it can answer
    questions directly — no code execution needed.
    """
    df = master_df.copy()

    # Ensure numeric columns
    for col in ["Debit", "Credit", "Balance", "Amount in USD (Reporting Currency)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Replace blank / NaN values in key groupby columns so the AI
    # doesn't see a mysterious "blank" entry as the largest group.
    for col in ["CostCenter", "Supplier", "Customer full name", "Mapping",
                "Distribution account", "Statement"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()
            df[col] = df[col].replace("", "Unassigned")

    num_cols = [
        c for c in ["Debit", "Credit", "Amount in USD (Reporting Currency)"]
        if c in df.columns
    ]

    ctx = []
    sep = "─" * 50

    ctx.append("═" * 60)
    ctx.append("LXT CONSOLIDATED FINANCIAL DATA")
    ctx.append("═" * 60)
    ctx.append(f"Total transactions: {len(df):,}")

    months = sorted(df["Reporting Month"].dropna().unique().tolist())
    if months:
        ctx.append(f"Reporting Months: {', '.join(str(m) for m in months)}")

    companies = sorted(df["Company Country"].dropna().unique().tolist())
    ctx.append(f"Companies ({len(companies)}): {', '.join(companies)}")

    # Grand totals
    ctx.append(f"\n{sep}")
    ctx.append("GRAND TOTALS")
    for col in num_cols:
        ctx.append(f"  {col}: ${df[col].sum():,.2f}")

    # By Company
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY COMPANY")
    ctx.append(df.groupby("Company Country")[num_cols].sum().round(2).to_string())

    # By Reporting Month
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY REPORTING MONTH")
    ctx.append(df.groupby("Reporting Month")[num_cols].sum().round(2).to_string())

    # By Company × Month
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY COMPANY × MONTH")
    ctx.append(
        df.groupby(["Company Country", "Reporting Month"])[num_cols]
        .sum().round(2).to_string()
    )

    # By Distribution Account
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY DISTRIBUTION ACCOUNT")
    ctx.append(df.groupby("Distribution account")[num_cols].sum().round(2).to_string())

    # By CostCenter
    if "CostCenter" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append("TOTALS BY COST CENTER")
        ctx.append(df.groupby("CostCenter")[num_cols].sum().round(2).to_string())

    # By Statement
    if "Statement" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append("TOTALS BY STATEMENT")
        ctx.append(
            df.groupby("Statement")["Amount in USD (Reporting Currency)"]
            .sum().round(2).to_string()
        )

    # By Mapping
    if "Mapping" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append("TOTALS BY MAPPING")
        ctx.append(
            df.groupby("Mapping")["Amount in USD (Reporting Currency)"]
            .sum().round(2).to_string()
        )

    # Company × Account
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY COMPANY × ACCOUNT")
    ctx.append(
        df.groupby(["Company Country", "Distribution account"])[num_cols]
        .sum().round(2).to_string()
    )

    # Company × CostCenter
    if "CostCenter" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append("TOTALS BY COMPANY × COST CENTER")
        ctx.append(
            df.groupby(["Company Country", "CostCenter"])[num_cols]
            .sum().round(2).to_string()
        )

    # Month × Account
    ctx.append(f"\n{sep}")
    ctx.append("TOTALS BY MONTH × ACCOUNT")
    ctx.append(
        df.groupby(["Reporting Month", "Distribution account"])[num_cols]
        .sum().round(2).to_string()
    )

    # Top Suppliers
    if "Supplier" in df.columns:
        sup = df[df["Supplier"].notna() & (df["Supplier"].astype(str).str.strip() != "")]
        if len(sup) > 0:
            ctx.append(f"\n{sep}")
            ctx.append("TOP 30 SUPPLIERS BY AMOUNT IN USD")
            top_sup = (
                sup.groupby("Supplier")["Amount in USD (Reporting Currency)"]
                .sum().round(2).sort_values(ascending=False).head(30)
            )
            ctx.append(top_sup.to_string())

    # Top Customers
    if "Customer full name" in df.columns:
        cust = df[
            df["Customer full name"].notna()
            & (df["Customer full name"].astype(str).str.strip() != "")
        ]
        if len(cust) > 0:
            ctx.append(f"\n{sep}")
            ctx.append("TOP 30 CUSTOMERS BY AMOUNT IN USD")
            top_cust = (
                cust.groupby("Customer full name")[
                    "Amount in USD (Reporting Currency)"
                ]
                .sum().round(2).sort_values(ascending=False).head(30)
            )
            ctx.append(top_cust.to_string())

    # Currency breakdown
    if "Currency" in df.columns:
        ctx.append(f"\n{sep}")
        ctx.append("TOTALS BY CURRENCY")
        ctx.append(df.groupby("Currency")[num_cols].sum().round(2).to_string())

    # Sample of raw General Ledger data (capped to stay within token limits)
    MAX_RAW_ROWS = 500
    ctx.append(f"\n{sep}")
    if len(df) <= MAX_RAW_ROWS:
        ctx.append(f"FULL GENERAL LEDGER DATA (CSV) — {len(df):,} rows")
        ctx.append(df.to_csv(index=False))
    else:
        ctx.append(
            f"GENERAL LEDGER SAMPLE ({MAX_RAW_ROWS} of {len(df):,} rows) "
            f"— use the pre-aggregated summaries above for accurate totals"
        )
        ctx.append(df.sample(n=MAX_RAW_ROWS, random_state=42).to_csv(index=False))

    # ── Exchange Rates ──
    if forex_rates:
        ctx.append(f"\n{'═' * 60}")
        ctx.append("EXCHANGE RATES (vs USD)")
        ctx.append("═" * 60)
        ctx.append("Each currency shows Closing Rate (for Balance Sheet items) "
                   "and Average Rate (for P&L items) per month.")
        ctx.append(f"Company → Currency: {', '.join(f'{k}: {v}' for k, v in COMPANY_CURRENCY.items())}")
        ctx.append("")
        for ccy in sorted(forex_rates.keys()):
            if ccy == "USD":
                continue
            ctx.append(f"  {ccy}:")
            months_data = forex_rates[ccy]
            for mk in sorted(months_data.keys()):
                rates = months_data[mk]
                ctx.append(
                    f"    {mk}: Closing = {rates['closing']:.6f}, "
                    f"Average = {rates['average']:.6f}"
                )
        ctx.append("  USD: All rates = 1.000000 (reporting currency)")

    # ── Account Mapping Reference ──
    if mapping_df is not None and not mapping_df.empty:
        ctx.append(f"\n{'═' * 60}")
        ctx.append("ACCOUNT MAPPING REFERENCE")
        ctx.append("═" * 60)
        ctx.append("This table shows how each Account Number is classified:")
        ctx.append("  - Mapping: the P&L or B/S category (e.g., Revenue, COGS, SG&A)")
        ctx.append("  - Item: whether it's P&L or Balance Sheet")
        ctx.append("  - Statement: the financial statement grouping")
        ctx.append("")
        ctx.append(mapping_df.to_string(index=False))

    # ── Pivot P&L ──
    if pivot_df is not None and len(pivot_df) > 0:
        ctx.append(f"\n{'═' * 60}")
        ctx.append("PIVOT P&L REPORT")
        ctx.append("═" * 60)
        ctx.append(pivot_df.to_string(index=False))

    return "\n".join(ctx)


# ═══════════════════════════════════════════════════════════════
# Financial Dashboard
# ═══════════════════════════════════════════════════════════════
def _render_financial_dashboard(master_df: pd.DataFrame) -> None:
    """Render an interactive financial analytics dashboard with Plotly charts."""
    import plotly.graph_objects as go
    import plotly.express as px

    df = master_df.copy()

    # Ensure numeric
    for col in ["Debit", "Credit", "Amount in USD (Reporting Currency)"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    pl = df[df["Item"].astype(str).str.strip().str.upper() == "P&L"]

    # ── Color palette (LXT brand) ──
    ORANGE = "#FE6F38"
    CORAL = "#E5592A"
    TEAL = "#00C9A7"
    RED = "#FF6B6B"
    BLUE = "#4ECDC4"
    PURPLE = "#A78BFA"
    AMBER = "#F59E0B"
    PINK = "#EC4899"
    LIGHT_BLUE = "#60A5FA"
    GRAY = "#64748B"

    chart_colors = [ORANGE, TEAL, BLUE, PURPLE, AMBER, PINK, CORAL, LIGHT_BLUE, RED, GRAY]

    # ── Plotly dark layout template ──
    dark_layout = dict(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, sans-serif", color="#C8D1DB", size=12),
        margin=dict(l=20, r=20, t=40, b=20),
        legend=dict(
            bgcolor="rgba(0,0,0,0)",
            bordercolor="rgba(255,255,255,0.05)",
            font=dict(size=11),
        ),
        xaxis=dict(
            gridcolor="rgba(255,255,255,0.04)",
            zerolinecolor="rgba(255,255,255,0.06)",
        ),
        yaxis=dict(
            gridcolor="rgba(255,255,255,0.04)",
            zerolinecolor="rgba(255,255,255,0.06)",
        ),
    )

    # ── Classification helper ──
    def _classify(statement: str) -> str:
        s = str(statement).strip().lower()
        if "revenue" in s:
            return "Revenue"
        elif any(k in s for k in ["cops", "direct cost", "inter-company cops"]):
            return "COPS"
        else:
            return "Expenses"

    pl = pl.copy()
    pl["_category"] = pl["Statement"].apply(_classify)
    amt_col = "Amount in USD (Reporting Currency)"

    # ── Compute KPIs ──
    total_revenue = pl[pl["_category"] == "Revenue"][amt_col].sum()
    total_cops = abs(pl[pl["_category"] == "COPS"][amt_col].sum())
    total_expenses = abs(pl[pl["_category"] == "Expenses"][amt_col].sum())
    gross_profit = total_revenue - total_cops
    gp_pct = (gross_profit / total_revenue * 100) if total_revenue != 0 else 0
    net_income = total_revenue - total_cops - total_expenses

    # ── KPI Cards ──
    def _kpi_html(label: str, value: str, css_class: str, sub: str = "") -> str:
        sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ""
        return (
            f'<div class="kpi-card">'
            f'<div class="kpi-label">{label}</div>'
            f'<div class="kpi-value {css_class}">{value}</div>'
            f'{sub_html}</div>'
        )

    # Row 1 of KPIs
    k1, k2, k3 = st.columns(3)
    with k1:
        st.markdown(_kpi_html("Total Revenue", f"${total_revenue:,.0f}", "positive"), unsafe_allow_html=True)
    with k2:
        st.markdown(_kpi_html("Cost of Services", f"${total_cops:,.0f}", "negative"), unsafe_allow_html=True)
    with k3:
        st.markdown(_kpi_html("Gross Profit", f"${gross_profit:,.0f}", "positive" if gross_profit > 0 else "negative"), unsafe_allow_html=True)

    # Row 2 of KPIs
    k4, k5, k6 = st.columns(3)
    with k4:
        st.markdown(_kpi_html("GP %", f"{gp_pct:.1f}%", "accent"), unsafe_allow_html=True)
    with k5:
        st.markdown(_kpi_html("Operating Expenses", f"${total_expenses:,.0f}", "negative"), unsafe_allow_html=True)
    with k6:
        st.markdown(_kpi_html("Net Income", f"${net_income:,.0f}", "positive" if net_income > 0 else "negative"), unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════
    # Row 1: Revenue by Company + Monthly P&L
    # ═══════════════════════════════════════════════════
    c1, c2 = st.columns(2)

    with c1:
        st.markdown('<div class="dashboard-section-title">💰 Revenue by Entity</div>', unsafe_allow_html=True)
        rev_data = pl[(pl["_category"] == "Revenue") & (pl["Company Country"].notna()) & (pl["Company Country"].astype(str).str.strip() != "")]
        rev_by_company = (
            rev_data
            .groupby("Company Country")[amt_col]
            .sum().round(2)
            .sort_values(ascending=True)
        )
        fig1 = go.Figure(go.Bar(
            x=rev_by_company.values,
            y=rev_by_company.index,
            orientation="h",
            marker=dict(
                color=rev_by_company.values,
                colorscale=[[0, CORAL], [0.5, ORANGE], [1, TEAL]],
                line=dict(width=0),
                cornerradius=4,
            ),
            text=[f"${v:,.0f}" for v in rev_by_company.values],
            textposition="auto",
            textfont=dict(size=11, color="white"),
        ))
        fig1.update_layout(**dark_layout, height=350, title="", showlegend=False)
        fig1.update_xaxes(title="Amount in USD")
        st.plotly_chart(fig1, width="stretch")

    with c2:
        st.markdown('<div class="dashboard-section-title">📊 Monthly P&L Breakdown</div>', unsafe_allow_html=True)
        # Sort months chronologically
        import calendar as _cal
        month_order = sorted(pl["Reporting Month"].dropna().unique().tolist())
        monthly = (
            pl.groupby(["Reporting Month", "_category"])[amt_col]
            .sum().round(2)
            .reset_index()
        )
        # Make COPS and Expenses positive for visual comparison
        monthly["_display_amt"] = monthly.apply(
            lambda r: abs(r[amt_col]) if r["_category"] != "Revenue" else r[amt_col], axis=1
        )
        cat_colors = {"Revenue": TEAL, "COPS": ORANGE, "Expenses": RED}
        fig2 = go.Figure()
        for cat in ["Revenue", "COPS", "Expenses"]:
            cat_data = monthly[monthly["_category"] == cat]
            fig2.add_trace(go.Bar(
                x=cat_data["Reporting Month"],
                y=cat_data["_display_amt"],
                name=cat,
                marker_color=cat_colors[cat],
                marker_cornerradius=4,
                text=[f"${v:,.0f}" for v in cat_data["_display_amt"]],
                textposition="outside",
                textfont=dict(size=10),
            ))
        fig2.update_layout(**dark_layout, height=350, barmode="group")
        fig2.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5))
        fig2.update_yaxes(title="Amount in USD")
        st.plotly_chart(fig2, width="stretch")

    # ═══════════════════════════════════════════════════
    # Row 2: Expense Breakdown + Cost Center Analysis
    # ═══════════════════════════════════════════════════
    c3, c4 = st.columns(2)

    with c3:
        st.markdown('<div class="dashboard-section-title">🍩 Expense Breakdown by Category</div>', unsafe_allow_html=True)
        expenses = pl[pl["_category"] == "Expenses"].copy()
        exp_by_stmt = (
            expenses.groupby("Statement")[amt_col]
            .sum().abs().round(2)
            .sort_values(ascending=False)
        )
        # Top 8 + Other
        if len(exp_by_stmt) > 8:
            top_8 = exp_by_stmt.head(8)
            other = pd.Series({"Other": exp_by_stmt.iloc[8:].sum()})
            exp_by_stmt = pd.concat([top_8, other])

        fig3 = go.Figure(go.Pie(
            labels=exp_by_stmt.index,
            values=exp_by_stmt.values,
            hole=0.55,
            marker=dict(colors=chart_colors[:len(exp_by_stmt)]),
            textinfo="percent",
            textfont=dict(size=11, color="white"),
            hovertemplate="%{label}<br>$%{value:,.2f}<br>%{percent}<extra></extra>",
        ))
        fig3.update_layout(**dark_layout, height=380, showlegend=True)
        fig3.update_layout(
            legend=dict(font=dict(size=10), x=0, y=-0.15, orientation="h"),
        )
        # Add center annotation
        fig3.add_annotation(
            text=f"<b>${exp_by_stmt.sum():,.0f}</b><br><span style='font-size:10px;color:#8899A6'>Total Expenses</span>",
            showarrow=False, font=dict(size=14, color="#F0F2F6"),
        )
        st.plotly_chart(fig3, width="stretch")

    with c4:
        st.markdown('<div class="dashboard-section-title">🏢 Cost Center Analysis</div>', unsafe_allow_html=True)
        cc_data = pl[pl["CostCenter"].notna() & (pl["CostCenter"].astype(str).str.strip() != "")]
        if len(cc_data) > 0:
            cc_summary = (
                cc_data.groupby("CostCenter")
                .agg(
                    Revenue=(amt_col, lambda x: x[cc_data.loc[x.index, "_category"] == "Revenue"].sum()),
                    Expenses=(amt_col, lambda x: abs(x[cc_data.loc[x.index, "_category"] != "Revenue"].sum())),
                )
                .round(2)
            )
            # Simpler approach: just total by cost center
            cc_totals = (
                cc_data.groupby("CostCenter")[amt_col]
                .sum().round(2)
                .sort_values(ascending=True)
            )
            colors = [TEAL if v > 0 else RED for v in cc_totals.values]
            fig4 = go.Figure(go.Bar(
                x=cc_totals.values,
                y=cc_totals.index,
                orientation="h",
                marker=dict(color=colors, cornerradius=4),
                text=[f"${v:,.0f}" for v in cc_totals.values],
                textposition="auto",
                textfont=dict(size=11, color="white"),
            ))
            fig4.update_layout(**dark_layout, height=380, showlegend=False)
            fig4.update_xaxes(title="Net Amount in USD")
            st.plotly_chart(fig4, width="stretch")
        else:
            st.info("No cost center data available.")

    # ═══════════════════════════════════════════════════
    # Row 3: Revenue Trend + Entity P&L Comparison
    # ═══════════════════════════════════════════════════
    c5, c6 = st.columns(2)

    with c5:
        st.markdown('<div class="dashboard-section-title">📈 Revenue Trend by Month</div>', unsafe_allow_html=True)
        rev_by_month_entity = (
            pl[pl["_category"] == "Revenue"]
            .groupby(["Reporting Month", "Company Country"])[amt_col]
            .sum().round(2)
            .reset_index()
        )
        fig5 = go.Figure()
        for i, company in enumerate(sorted(rev_by_month_entity["Company Country"].unique())):
            comp_data = rev_by_month_entity[rev_by_month_entity["Company Country"] == company]
            fig5.add_trace(go.Scatter(
                x=comp_data["Reporting Month"],
                y=comp_data[amt_col],
                mode="lines+markers",
                name=company,
                line=dict(width=2.5, color=chart_colors[i % len(chart_colors)]),
                marker=dict(size=8),
            ))
        fig5.update_layout(**dark_layout, height=380)
        fig5.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5, font=dict(size=9)))
        fig5.update_yaxes(title="Revenue (USD)")
        st.plotly_chart(fig5, width="stretch")

    with c6:
        st.markdown('<div class="dashboard-section-title">🏦 Entity P&L Comparison</div>', unsafe_allow_html=True)
        entity_pl = (
            pl.groupby(["Company Country", "_category"])[amt_col]
            .sum().round(2)
            .reset_index()
        )
        entity_pl["_display_amt"] = entity_pl.apply(
            lambda r: abs(r[amt_col]) if r["_category"] != "Revenue" else r[amt_col], axis=1
        )
        companies = sorted(entity_pl["Company Country"].unique())
        fig6 = go.Figure()
        for cat, color in [("Revenue", TEAL), ("COPS", ORANGE), ("Expenses", RED)]:
            cat_data = entity_pl[entity_pl["_category"] == cat].set_index("Company Country").reindex(companies).fillna(0)
            fig6.add_trace(go.Bar(
                x=companies,
                y=cat_data["_display_amt"].values if "_display_amt" in cat_data.columns else [0]*len(companies),
                name=cat,
                marker_color=color,
                marker_cornerradius=4,
            ))
        fig6.update_layout(**dark_layout, height=380, barmode="group")
        fig6.update_layout(
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
            xaxis=dict(tickangle=-30, tickfont=dict(size=10)),
        )
        fig6.update_yaxes(title="Amount (USD)")
        st.plotly_chart(fig6, width="stretch")


def _run_etl(start_date: str, end_date: str, forex_rates: dict, mapping_df: pd.DataFrame):
    """Execute the full ETL pipeline with progress UI."""

    # Clear cached AI and pivot context so they rebuild with new data
    st.session_state.pop("financial_context", None)
    st.session_state.pop("messages", None)
    st.session_state.pop("pivot_data", None)
    st.session_state.pop("pivot_name", None)
    st.session_state.pop("pivot_preview", None)
    st.session_state.pop("pivot_rows", None)

    # Load credentials from secrets / environment variables
    client_id = get_secret("QB_CLIENT_ID")
    client_secret = get_secret("QB_CLIENT_SECRET")
    companies = _load_companies()

    # Load latest tokens from GitHub Gist (falls back to secrets.toml)
    gist_tokens = _load_gist_tokens() or {}
    updated_tokens: dict[str, str] = {}

    all_frames: list[pd.DataFrame] = []
    errors: list[str] = []

    with st.spinner("Fetching data from QuickBooks…"):
        progress = st.progress(0, text="Starting…")
        status_container = st.status(
            "Processing 9 companies…", expanded=True
        )

        company_keys = list(companies.keys())
        total = len(company_keys)

        for idx, key in enumerate(company_keys):
            company = companies[key]
            label = company["label"]
            realm_id = company["realm_id"]
            # Prefer gist token, fall back to secrets.toml
            refresh_token = gist_tokens.get(key, company["refresh_token"])

            progress.progress(
                (idx) / total,
                text=f"Processing {label} ({idx + 1}/{total})…",
            )

            try:
                with status_container:
                    st.write(f"🔄 **{label}** — Authenticating…")

                # Auth — try gist token first, fall back to secrets.toml
                token_info = None
                secrets_token = company["refresh_token"]
                used_token = refresh_token
                try:
                    token_info = refresh_access_token(
                        client_id, client_secret, refresh_token
                    )
                except RuntimeError as auth_err:
                    # If gist token failed and we have a different secrets.toml token, retry
                    if (
                        "invalid_grant" in str(auth_err)
                        and secrets_token
                        and secrets_token != refresh_token
                    ):
                        with status_container:
                            st.write(
                                f"⚠️ **{label}** — Gist token expired, "
                                f"retrying with secrets.toml token…"
                            )
                        token_info = refresh_access_token(
                            client_id, client_secret, secrets_token
                        )
                        used_token = secrets_token
                    else:
                        raise

                # Track the new refresh token
                new_refresh = token_info["refresh_token"]
                updated_tokens[key] = new_refresh

                # Also save locally (best-effort)
                if new_refresh != used_token:
                    _save_refresh_token(used_token, new_refresh)

                with status_container:
                    st.write(f"📥 **{label}** — Fetching General Ledger…")

                # Extract
                report_json = fetch_general_ledger(
                    token_info["access_token"], realm_id, start_date, end_date
                )
                raw_rows = flatten_report(report_json)

                # Transform
                df = transform(raw_rows, label)

                with status_container:
                    st.write(f"✅ **{label}** — {len(df)} rows extracted.")

                if not df.empty:
                    all_frames.append(df)

            except Exception as exc:
                msg = f"❌ **{label}** — {exc}"
                errors.append(msg)
                with status_container:
                    st.write(msg)

        progress.progress(1.0, text="Done!")
        status_container.update(
            label="Processing complete!", state="complete", expanded=False
        )

    # ── Save updated tokens to GitHub Gist ─────────────────────
    if updated_tokens:
        _save_gist_tokens(updated_tokens)

    # ── Results ───────────────────────────────────────────────
    st.divider()

    if errors:
        with st.expander(f"⚠️ {len(errors)} error(s) occurred", expanded=False):
            for e in errors:
                st.markdown(e)

    if not all_frames:
        st.warning("No data was collected from any company.")
        return

    master_df = pd.concat(all_frames, ignore_index=True)

    # Apply Consol Mapping Sheet lookup
    master_df = apply_mapping(master_df, mapping_df)

    # Apply Forex Rate based on Currency + Reporting Month + Item
    def _get_forex_rate(row):
        ccy = str(row.get("Currency", "")).strip().upper()
        item = str(row.get("Item", "")).strip()
        month = str(row.get("Reporting Month", "")).strip()

        # USD is always 1.0
        if ccy == "USD":
            return 1.0

        # Look up per-month rates for the currency
        month_rates = forex_rates.get(ccy, {})
        rates = month_rates.get(month, {"closing": 1.0, "average": 1.0})

        if item == "P&L":
            return rates["average"]
        else:
            return rates["closing"]

    master_df["Forex Rate"] = master_df.apply(_get_forex_rate, axis=1)
    master_df["Amount in USD (Reporting Currency)"] = (
        master_df["Transaction Value in Original Currency"] * master_df["Forex Rate"]
    )

    # Build Excel bytes and filename
    xlsx_bytes = to_excel_bytes(master_df)
    sd = datetime.strptime(start_date, "%Y-%m-%d")
    ed = datetime.strptime(end_date, "%Y-%m-%d")
    file_name = f"LXT_General_Ledger_{sd.strftime('%d%b%Y')}_to_{ed.strftime('%d%b%Y')}.xlsx"

    # Store in session state so the download button persists
    st.session_state["report_data"] = xlsx_bytes
    st.session_state["report_name"] = file_name
    st.session_state["report_rows"] = len(master_df)
    st.session_state["report_preview"] = master_df.head(100)

    # Store master_df for pivot report generation
    st.session_state["master_df"] = master_df

    # Store forex rates and mapping for AI context
    st.session_state["forex_rates"] = forex_rates
    st.session_state["mapping_df"] = mapping_df


# ═══════════════════════════════════════════════════════════════
# Entry Point
# ═══════════════════════════════════════════════════════════════
if check_password():
    main_app()
